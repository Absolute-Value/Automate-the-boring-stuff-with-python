#! /usr/bin/env python3
# 12.14.1_multiplicationTable.py - 掛け算の表を作成する

import sys, openpyxl
from openpyxl.styles import Font

def multiplication_table(num):
    wb = openpyxl.Workbook()
    sheet = wb.active

    for i in range(1, num+1):
        cell = sheet.cell(1,i+1)
        cell.value = i
        cell.font = Font(bold=True)

        for j in range(1, num+1):
            cell = sheet.cell(j+1,i+1)
            cell.value = j * i

    for j in range(1, num+1):
        cell = sheet.cell(j+1,1)
        cell.value = j
        cell.font = Font(bold=True)

    wb.save('MultiplicationTable.xlsx')

if len(sys.argv) < 2:
    print('数字を入力してください')
else:
    multiplication_table(int(sys.argv[1]))